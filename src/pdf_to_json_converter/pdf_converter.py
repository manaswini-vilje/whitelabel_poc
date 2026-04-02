"""
Document to JSON Converter Module
Efficiently converts documents (PDF, images, etc.) to JSON using Azure Document Intelligence + GPT-4.
"""

import os
import json
import re
import time
from pathlib import Path
from typing import Optional, Dict, Any, List
from datetime import datetime, date
from dotenv import load_dotenv
from openai import AzureOpenAI
from openai import APITimeoutError, APIError, RateLimitError
from azure.core.credentials import AzureKeyCredential
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.exceptions import AzureError

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Load environment variables
load_dotenv()


class DocumentToJSONConverter:
    """Efficiently converts documents to JSON using Azure Document Intelligence + GPT-4."""
    
    def __init__(self, api_key: Optional[str] = None, azure_endpoint: Optional[str] = None, azure_key: Optional[str] = None):
        """Initialize the converter."""
        # Azure OpenAI client
        self.azure_openai_endpoint = azure_endpoint or os.getenv("AZURE_OPENAI_ENDPOINT")
        self.azure_openai_api_key = api_key or os.getenv("AZURE_OPENAI_API_KEY")
        self.azure_openai_deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4o-mini")
        self.azure_openai_api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2025-01-01-preview")
        
        if not self.azure_openai_endpoint or not self.azure_openai_api_key:
            raise ValueError(
                "Azure OpenAI credentials are required. "
                "Set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY environment variables."
            )
        
        self.client = AzureOpenAI(
            api_key=self.azure_openai_api_key,
            api_version=self.azure_openai_api_version,
            azure_endpoint=self.azure_openai_endpoint,
            timeout=300.0,  # Increased timeout for large responses (5 minutes)
            max_retries=2
        )
        
        # Azure Form Recognizer client
        self.azure_endpoint = azure_endpoint or os.getenv("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT")
        self.azure_key = azure_key or os.getenv("AZURE_DOCUMENT_INTELLIGENCE_KEY")
        
        if not self.azure_endpoint or not self.azure_key:
            raise ValueError(
                "Azure Form Recognizer credentials are required. "
                "Set AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT and AZURE_DOCUMENT_INTELLIGENCE_KEY environment variables."
            )
        
        self.azure_client = DocumentAnalysisClient(
            endpoint=self.azure_endpoint,
            credential=AzureKeyCredential(self.azure_key)
        )
    
    def _is_supported_format(self, file_path: str) -> bool:
        """Check if file format is directly supported by Azure Document Intelligence."""
        supported_extensions = {'.pdf', '.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.docx', '.doc', '.pptx', '.ppt'}
        file_ext = Path(file_path).suffix.lower()
        return file_ext in supported_extensions
    
    def _extract_excel_data(self, excel_path: str) -> Dict[str, Any]:
        """Extract data directly from Excel file without converting to PDF."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel processing. Install with: pip install openpyxl")
        
        try:
            print(f"  Reading Excel file directly...")
            
            # Read Excel file
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            
            all_tables = []
            text_content = []
            total_sheets = len(workbook.sheetnames)
            
            # Process each worksheet
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames, 1):
                sheet = workbook[sheet_name]
                print(f"  Processing sheet {sheet_idx}/{total_sheets}: {sheet_name}")
                
                # Get all data from sheet
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                if max_row == 0 or max_col == 0:
                    print(f"    ⚠️  Sheet '{sheet_name}' is empty, skipping...")
                    continue
                
                # Extract all data including empty cells to preserve structure
                sheet_data = []
                for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False):
                    row_data = []
                    for cell in row:
                        # Get cell value, handling different types
                        if cell.value is None:
                            cell_str = ""
                        elif isinstance(cell.value, (int, float)):
                            # Preserve numeric values as strings but keep precision
                            if isinstance(cell.value, float) and cell.value.is_integer():
                                cell_str = str(int(cell.value))
                            else:
                                cell_str = str(cell.value)
                        elif isinstance(cell.value, (datetime, date)):
                            # Format dates consistently
                            if isinstance(cell.value, datetime):
                                cell_str = cell.value.strftime("%Y-%m-%d %H:%M:%S")
                            else:  # date object
                                cell_str = cell.value.strftime("%Y-%m-%d")
                        else:
                            cell_str = str(cell.value)
                        row_data.append(cell_str)
                    sheet_data.append(row_data)
                
                if sheet_data:
                    # Add sheet information to text
                    text_content.append(f"=== Sheet {sheet_idx}: {sheet_name} ===")
                    text_content.append(f"Total rows: {max_row}, Total columns: {max_col}\n")
                    
                    # Format as table for text representation
                    if len(sheet_data) > 0:
                        # Use first row as potential headers
                        headers = sheet_data[0] if sheet_data else []
                        text_content.append("Headers: " + " | ".join(str(h) if h else "" for h in headers))
                        text_content.append("\nData rows:")
                        
                        # Add all rows to text
                        for row_idx, row in enumerate(sheet_data[1:] if len(sheet_data) > 1 else sheet_data, start=1):
                            row_str = " | ".join(str(cell) if cell else "" for cell in row)
                            text_content.append(f"  Row {row_idx}: {row_str}")
                    
                    text_content.append("\n")
                    
                    # Store as table structure
                    all_tables.append({
                        'page': sheet_idx,  # Use sheet index as page number
                        'table_num': len(all_tables) + 1,
                        'data': sheet_data,
                        'row_count': len(sheet_data),
                        'sheet_name': sheet_name
                    })
                    
                    print(f"    ✓ Extracted {len(sheet_data)} row(s) with {max_col} column(s)")
            
            # Combine text content
            combined_text = '\n'.join(text_content)
            
            print(f"  ✓ Excel extraction completed: {len(all_tables)} table(s) from {total_sheets} sheet(s)")
            
            return {
                'text': combined_text,
                'tables': all_tables,
                'total_pages': total_sheets  # Use sheet count as page count
            }
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def _is_excel_file(self, file_path: str) -> bool:
        """Check if file is an Excel file."""
        file_ext = Path(file_path).suffix.lower()
        return file_ext in {'.xlsx', '.xls'}
    
    def _split_pdf_into_chunks(self, pdf_path: str, chunk_size: int = 2) -> List[str]:
        """Split PDF into chunks of specified page size using PyMuPDF."""
        if not PYMUPDF_AVAILABLE:
            raise ImportError("PyMuPDF (fitz) is required for chunk processing. Install with: pip install PyMuPDF")
        
        try:
            src = fitz.open(pdf_path)
            total_pages = len(src)
            chunk_files = []
            
            temp_dir = Path(pdf_path).parent / "temp_chunks"
            temp_dir.mkdir(exist_ok=True)
            
            print(f"  Splitting PDF into chunks of {chunk_size} pages...")
            print(f"  Total pages: {total_pages}")
            
            for start_page in range(0, total_pages, chunk_size):
                end_page = min(start_page + chunk_size, total_pages)
                chunk_pdf = str(temp_dir / f"chunk_{start_page + 1}_{end_page}.pdf")
                
                dst = fitz.open()
                for page_num in range(start_page, end_page):
                    page = src[page_num]
                    new_page = dst.new_page(width=page.rect.width, height=page.rect.height)
                    new_page.show_pdf_page(new_page.rect, src, page_num)
                
                dst.save(chunk_pdf)
                dst.close()
                chunk_files.append(chunk_pdf)
                print(f"    Created chunk: pages {start_page + 1}-{end_page}")
            
            src.close()
            print(f"  ✓ Created {len(chunk_files)} chunk(s)")
            return chunk_files
            
        except Exception as e:
            print(f"❌ Error splitting PDF: {e}")
            raise
    
    def _extract_chunk_with_azure(self, chunk_path: str, chunk_num: int, total_chunks: int) -> Any:
        """Extract text and tables from a single chunk using Azure Document Intelligence."""
        print(f"  Processing chunk {chunk_num}/{total_chunks}...")
        try:
            with open(chunk_path, "rb") as f:
                poller = self.azure_client.begin_analyze_document(
                    model_id="prebuilt-document",
                    document=f
                )
            
            result = poller.result()
            print(f"  ✓ Chunk {chunk_num} analysis completed")
            return result
            
        except Exception as e:
            print(f"  ❌ Error processing chunk {chunk_num}: {e}")
            raise
    
    def _extract_document_with_azure(self, document_path: str) -> Dict[str, Any]:
        """Extract text and tables from document using Azure Document Intelligence - processes in 2-page chunks for PDFs."""
        print("Extracting text and tables from document using Azure Document Intelligence...")
        
        doc_path = Path(document_path)
        is_pdf = doc_path.suffix.lower() == '.pdf'
        
        # For PDFs, process in 2-page chunks to ensure all pages are processed
        if is_pdf and PYMUPDF_AVAILABLE:
            try:
                # Split PDF into 2-page chunks
                print("\n[Step 1/2] Splitting PDF into 2-page chunks...")
                chunk_files = self._split_pdf_into_chunks(document_path, chunk_size=2)
                total_chunks = len(chunk_files)
                
                # Process each chunk
                print(f"\n[Step 2/2] Processing {total_chunks} chunk(s) with Azure Document Intelligence...")
                chunk_results = []
                
                try:
                    for idx, chunk_file in enumerate(chunk_files, 1):
                        result = self._extract_chunk_with_azure(chunk_file, idx, total_chunks)
                        chunk_results.append({
                            'result': result,
                            'chunk_num': idx
                        })
                    
                    # Combine results from all chunks
                    print(f"\n[Step 3/3] Combining results from all chunks...")
                    pages_text_dict = {}
                    all_tables = []
                    
                    for chunk_info in chunk_results:
                        result = chunk_info['result']
                        chunk_num = chunk_info['chunk_num']
                        page_offset = (chunk_num - 1) * 2
                        
                        # Extract text from pages
                        if result.pages:
                            for page in result.pages:
                                actual_page_num = page_offset + page.page_number
                                page_text_lines = []
                                if hasattr(page, 'lines') and page.lines:
                                    for line in page.lines:
                                        if line.content and line.content.strip():
                                            page_text_lines.append(line.content)
                                if page_text_lines:
                                    if actual_page_num not in pages_text_dict:
                                        pages_text_dict[actual_page_num] = []
                                    pages_text_dict[actual_page_num].extend(page_text_lines)
                        
                        # Extract paragraphs
                        if hasattr(result, 'paragraphs') and result.paragraphs:
                            for para in result.paragraphs:
                                if para.bounding_regions and para.content and para.content.strip():
                                    actual_page_num = page_offset + para.bounding_regions[0].page_number
                                    if actual_page_num not in pages_text_dict:
                                        pages_text_dict[actual_page_num] = []
                                    para_text = para.content.strip()
                                    existing_text = '\n'.join(pages_text_dict[actual_page_num])
                                    if para_text not in existing_text:
                                        pages_text_dict[actual_page_num].append(para_text)
                        
                        # Extract tables
                        if result.tables:
                            for table in result.tables:
                                if table.bounding_regions:
                                    actual_page_num = page_offset + table.bounding_regions[0].page_number
                                    
                                    table_data = []
                                    if table.cells:
                                        rows = {}
                                        for cell in table.cells:
                                            row_idx = cell.row_index
                                            col_idx = cell.column_index
                                            if row_idx not in rows:
                                                rows[row_idx] = {}
                                            rows[row_idx][col_idx] = cell.content if cell.content else ""
                                        
                                        if rows:
                                            max_row = max(rows.keys())
                                            max_col = max(max(rows[r].keys()) for r in rows if rows[r])
                                            for row_idx in range(max_row + 1):
                                                row = []
                                                for col_idx in range(max_col + 1):
                                                    cell_content = rows.get(row_idx, {}).get(col_idx, "")
                                                    row.append(cell_content)
                                                table_data.append(row)
                                    
                                    all_tables.append({
                                        'page': actual_page_num,
                                        'table_num': len(all_tables) + 1,
                                        'data': table_data,
                                        'row_count': len(table_data)
                                    })
                    
                    # Format text by pages
                    text_content = []
                    for page_num in sorted(pages_text_dict.keys()):
                        page_text = '\n'.join(pages_text_dict[page_num])
                        text_content.append(f"=== Page {page_num} ===\n{page_text}\n")
                    
                    combined_text = '\n'.join(text_content)
                    
                    # Determine total pages
                    total_pages = max(pages_text_dict.keys()) if pages_text_dict else 1
                    if all_tables:
                        max_table_page = max([t['page'] for t in all_tables])
                        total_pages = max(total_pages, max_table_page)
                    
                    # Cleanup chunk files
                    print(f"\n  Cleaning up temporary chunk files...")
                    for chunk_file in chunk_files:
                        try:
                            os.remove(chunk_file)
                        except:
                            pass
                    try:
                        temp_dir = Path(chunk_files[0]).parent
                        if temp_dir.exists():
                            temp_dir.rmdir()
                    except:
                        pass
                    print(f"  ✓ Cleanup completed")
                    
                    return {
                        'text': combined_text,
                        'tables': all_tables,
                        'total_pages': total_pages
                    }
                    
                except Exception as e:
                    # Cleanup on error
                    for chunk_file in chunk_files:
                        try:
                            os.remove(chunk_file)
                        except:
                            pass
                    raise
                    
            except Exception as e:
                print(f"⚠️  Chunking failed: {e}")
                print(f"⚠️  Falling back to direct processing...")
                # Fall through to direct processing
        
        # Direct processing (for non-PDFs or if chunking fails)
        try:
            with open(document_path, "rb") as f:
                poller = self.azure_client.begin_analyze_document(
                    model_id="prebuilt-document",
                    document=f
                )
            
            print("  Waiting for analysis to complete...")
            result = poller.result()
            print("  ✓ Analysis completed")
            
            # Extract text content
            text_content = []
            pages_text_dict = {}
            
            # Extract from pages
            if result.pages:
                for page in result.pages:
                    page_num = page.page_number
                    page_text_lines = []
                    if hasattr(page, 'lines') and page.lines:
                        for line in page.lines:
                            if line.content and line.content.strip():
                                page_text_lines.append(line.content)
                    if page_text_lines:
                        pages_text_dict[page_num] = page_text_lines
            
            # Extract paragraphs
            if hasattr(result, 'paragraphs') and result.paragraphs:
                for para in result.paragraphs:
                    if para.bounding_regions and para.content and para.content.strip():
                        page_num = para.bounding_regions[0].page_number
                        if page_num not in pages_text_dict:
                            pages_text_dict[page_num] = []
                        para_text = para.content.strip()
                        existing_text = '\n'.join(pages_text_dict[page_num])
                        if para_text not in existing_text:
                            pages_text_dict[page_num].append(para_text)
            
            # Format text by pages
            for page_num in sorted(pages_text_dict.keys()):
                page_text = '\n'.join(pages_text_dict[page_num])
                text_content.append(f"=== Page {page_num} ===\n{page_text}\n")
            
            combined_text = '\n'.join(text_content)
            
            # Extract tables
            tables_data = []
            if result.tables:
                for table_idx, table in enumerate(result.tables, 1):
                    table_data = []
                    if table.cells:
                        rows = {}
                        for cell in table.cells:
                            row_idx = cell.row_index
                            col_idx = cell.column_index
                            if row_idx not in rows:
                                rows[row_idx] = {}
                            rows[row_idx][col_idx] = cell.content if cell.content else ""
                        
                        if rows:
                            max_row = max(rows.keys())
                            max_col = max(max(rows[r].keys()) for r in rows if rows[r])
                            for row_idx in range(max_row + 1):
                                row = []
                                for col_idx in range(max_col + 1):
                                    cell_content = rows.get(row_idx, {}).get(col_idx, "")
                                    row.append(cell_content)
                                table_data.append(row)
                    
                    page_num = table.bounding_regions[0].page_number if table.bounding_regions else 1
                    tables_data.append({
                        'page': page_num,
                        'table_num': table_idx,
                        'data': table_data,
                        'row_count': len(table_data)
                    })
            
            # Determine total pages
            total_pages = len(result.pages) if result.pages else 1
            if tables_data:
                max_table_page = max([t['page'] for t in tables_data])
                total_pages = max(total_pages, max_table_page)
            
            return {
                'text': combined_text,
                'tables': tables_data,
                'total_pages': total_pages
            }
            
        except AzureError as e:
            print(f"❌ Azure Document Intelligence error: {e}")
            raise Exception(f"Failed to extract text from document: {e}")
        except Exception as e:
            print(f"❌ Error extracting document: {e}")
            raise
    
    def _format_tables_for_prompt(self, tables: List[Dict], start_row: int = 0, end_row: int = None) -> str:
        """Format extracted tables for the prompt - simplified. Can format a subset of rows."""
        if not tables:
            return ""
        
        formatted = "\n\n=== EXTRACTED TABLES (EXTRACT ALL ROWS FROM ALL PAGES/SHEETS) ===\n"
        total_rows = 0
        rows_included = 0
        
        for table_info in tables:
            table = table_info['data']
            row_count = table_info.get('row_count', len(table))
            total_rows += row_count
            page_or_sheet = table_info.get('sheet_name', f"Page {table_info['page']}")
            
            if table:
                # Use first row as headers if available
                headers = table[0] if table else []
                
                # Determine which rows to include
                data_rows = table[1:] if len(table) > 1 else table
                
                if end_row is not None:
                    # Include only rows in the specified range
                    chunk_rows = data_rows[start_row:end_row]
                    if chunk_rows:
                        formatted += f"\nTable {table_info['table_num']} ({page_or_sheet}) - Rows {start_row + 1} to {min(end_row, len(data_rows))} of {row_count}:\n"
                        formatted += "Headers: " + " | ".join(str(h) if h else "" for h in headers) + "\n"
                        formatted += "Data Rows:\n"
                        for idx, row in enumerate(chunk_rows, start=start_row + 1):
                            formatted += f"  Row {idx}: " + " | ".join(str(cell) if cell else "" for cell in row) + "\n"
                        rows_included += len(chunk_rows)
                else:
                    # Include all rows
                    formatted += f"\nTable {table_info['table_num']} ({page_or_sheet}) - {row_count} ROWS:\n"
                    formatted += "Headers: " + " | ".join(str(h) if h else "" for h in headers) + "\n"
                    formatted += "Data Rows:\n"
                    for idx, row in enumerate(data_rows, start=1):
                        formatted += f"  Row {idx}: " + " | ".join(str(cell) if cell else "" for cell in row) + "\n"
                    rows_included += len(data_rows)
        
        if end_row is not None:
            formatted += f"\nCHUNK: Processing rows {start_row + 1} to {end_row} (Total: {total_rows} rows in document).\n"
        else:
            formatted += f"\nTOTAL: {total_rows} rows. YOU MUST EXTRACT ALL {total_rows} ROWS FROM ALL PAGES/SHEETS.\n"
        return formatted
    
    def _chunk_tables(self, tables: List[Dict], chunk_size: int = 150) -> List[tuple]:
        """Split tables into chunks for processing. Returns list of (start_row, end_row) tuples."""
        if not tables:
            return [(0, None)]
        
        # Calculate total data rows (excluding headers)
        total_data_rows = 0
        for table_info in tables:
            table = table_info.get('data', [])
            # Subtract 1 for header row if present
            data_rows = len(table) - 1 if len(table) > 1 else len(table)
            total_data_rows += data_rows
        
        if total_data_rows <= chunk_size:
            return [(0, None)]  # Process all at once
        
        # Create chunks
        chunks = []
        for start in range(0, total_data_rows, chunk_size):
            end = min(start + chunk_size, total_data_rows)
            chunks.append((start, end))
        
        return chunks
    
    def _call_gpt4_api(self, text_content: str, tables_content: str = "", retry_count: int = 0, total_pages: int = 1, chunk_info: str = "") -> str:
        """Call GPT-4 API to convert extracted text to JSON."""
        chunk_note = f"\n⚠️ IMPORTANT: This is processing {chunk_info} of the document. " if chunk_info else ""
        chunk_instruction = "Extract ONLY the products from the SPECIFIC ROWS indicated in the tables section below (look for 'CHUNK: Processing rows X to Y'). Do NOT extract products from other rows." if chunk_info else "Extract ALL products from ALL pages."

        prompt = f"""Analyze the following document text and extract information into a well-structured JSON format.
{chunk_note}
This document has {total_pages} pages. {chunk_instruction}

CRITICAL REQUIREMENTS - EXACT JSON STRUCTURE REQUIRED:
1. Extract EVERY SINGLE ROW from the tables section below - do not skip any rows
   - This document has multiple pages - you MUST process ALL pages including Page 1, Page 2, Page 3, and ALL subsequent pages
   - Count the rows in the tables section below - you MUST extract that exact number of product rows
   - DO NOT skip rows, summarize, or truncate - extract EVERY row as a separate product entry
   - CRITICAL: Make sure you extract data from Page 3 and all other pages - check the "=== Page X ===" sections
2. For fields containing multiple values separated by commas (e.g., "Z/39/11/2025/TRX, Order: 15001208177, EAN: 5903211190728"), 
   extract them as SEPARATE key-value pairs:
   - batch_number: Extract the batch/reference number (e.g., "Z/39/11/2025/TRX")
   - order_number: Extract the order number (e.g., "15001208177")  
   - ean: Extract the EAN code (e.g., "5903211190728")
3. Include ALL table rows from ALL pages - do not summarize, truncate, or skip any rows
   - Verify: The number of products in your output must match the number of data rows in the tables
   - Make sure you check Page 1, Page 2, Page 3, and all other pages for table data
4. Preserve all numeric values, dates, and text exactly as they appear
5. If a row appears empty or has missing data, still include it with empty strings for missing fields

REQUIRED JSON STRUCTURE - ADAPT TO DOCUMENT STRUCTURE:
Analyze the document structure FIRST, then extract data using appropriate field names based on what columns/fields are actually present in the document.

Base structure:
{{
  "document_metadata": {{
    "document_type": "Packing List" or "Invoice" or "Faktura" etc. (use what's in document),
    "document_number": "document number from header (could be FA/XXX, invoice number, etc.)",
    "date_of_issue": "date in DD-MM-YYYY format",
    "seller": {{
      "name": "seller company name",
      "address": "full address",
      "NIP": "NIP number if present",
      "REGON": "REGON number if present",
      "phone": "phone number (use 'phone' not 'telephone')",
      "BDO": "BDO number if present"
    }},
    "payer": {{
      "name": "payer company name",
      "address": "full address",
      "EURO_NIP": "EURO NIP if present"
    }},
    "buyer": {{
      "name": "buyer company name",
      "address": "full address",
      "delivery_address": "delivery address if different"
    }}
  }},
  "products": [
    {{
      "product_service_name": "product name (extract from product/service column)",
      "fabric_colors": "fabric/color description (extract from fabric/color column if present)",
      "batch_number": "batch number (extract from batch/reference column)",
      "order_number": "order number (extract from order column if present)",
      "ean": "EAN code (extract from EAN column if present)",
      "article": "article/SKU code (extract from article/SKU column - DO NOT leave empty if present in document)",
      "site": "site/location if present (extract from site column if present, otherwise empty string)",
      "quantity": numeric_value (extract from quantity column),
      "gross_quantity": numeric_value (extract from gross quantity column if present, otherwise same as quantity),
      "net_weight": numeric_value (extract from net weight column if present, 0 if not),
      "gross_weight": numeric_value (extract from gross weight column if present, 0 if not)
    }}
  ],
  "totals": {{
    "total_quantity": sum_of_all_quantities,
    "total_gross_quantity": sum_of_all_gross_quantities,
    "total_net_weight": sum_of_all_net_weights,
    "total_gross_weight": sum_of_all_gross_weights
  }}
}}

FIELD EXTRACTION RULES - ADAPT TO ACTUAL COLUMNS:
1. Look at the table headers to identify what columns are present
2. Map columns to fields based on their content:
   - Product name column → "product_service_name"
   - Fabric/Color/Description column → "fabric_colors"
   - Batch/Reference column → "batch_number"
   - Order column → "order_number"
   - EAN/Barcode column → "ean"
   - Article/SKU/Item code column → "article" (CRITICAL: extract this if present, don't leave empty)
   - Site/Location column → "site"
   - Quantity column → "quantity"
   - Gross quantity column → "gross_quantity"
   - Net weight column → "net_weight"
   - Gross weight column → "gross_weight"
3. If a column contains multiple pieces of information (e.g., "BASE:POSO 100 SEAT:POSO 100"), put it in "fabric_colors"
4. If article/SKU is missing from a row but present in other rows, check if it's in a different column or combined with another field
5. DO NOT skip the "article" field - if it exists in the document, extract it for every row

IMPORTANT:
- Use "products" array (NOT "tables" or "items")
- Use "buyer" in metadata (NOT "recipient")
- Use "phone" in seller (NOT "telephone")
- Use "document_number" in metadata (adapt to what's in document: could be "FA/XXX", invoice number, etc.)
{f"- Extract document_metadata ONLY if this is the first chunk - otherwise use empty object {{}}" if chunk_info else "- Calculate totals section by summing all product values"}
{f"- Do NOT calculate totals in this chunk - totals will be calculated after merging all chunks" if chunk_info else "- Include ALL products from ALL pages (Page 1, Page 2, Page 3, and beyond)"}
- If article code is in the document, extract it - don't leave it empty

{tables_content}

=== DOCUMENT TEXT ===
{text_content}

Return ONLY valid JSON matching the exact structure above. Do not include any explanations or markdown formatting, just the JSON object."""

        print("Calling GPT-4 API to structure data...")
        start_time = time.time()
        
        try:
            # Calculate approximate token count (rough estimate: 1 token ≈ 4 characters)
            prompt_length = len(prompt)
            estimated_tokens = prompt_length // 4
            print(f"  Prompt size: ~{estimated_tokens:,} tokens ({prompt_length:,} characters)")
            
            # Use maximum tokens for large documents to avoid truncation
            # GPT-4o supports up to 16,384 output tokens (this is the maximum)
            # For outputs exceeding this, we'll use continuation mechanism
            max_tokens = 16384  # Maximum for GPT-4o
            model_name = self.azure_openai_deployment  # Use deployment name for Azure OpenAI
            
            all_responses = []
            continuation_count = 0
            max_continuations = 5  # Allow up to 5 continuations (total ~80k tokens)
            
            # Initial request
            response = self.client.chat.completions.create(
                model=model_name,
                messages=[
                    {
                        "role": "system",
                        "content": "You are a data extraction expert. Extract ALL information from ALL pages of documents and return ONLY valid, complete JSON. The JSON must be complete and properly closed. Never truncate or cut off the response. If the data is too large, ensure the JSON structure is complete with all closing brackets."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                max_tokens=max_tokens,
                temperature=0.1
            )
            
            response_text = response.choices[0].message.content
            finish_reason = response.choices[0].finish_reason
            all_responses.append(response_text)
            
            # Handle continuations if response was truncated
            while finish_reason == "length" and continuation_count < max_continuations:
                continuation_count += 1
                print(f"⚠️  Response was truncated (continuation {continuation_count}/{max_continuations})")
                print(f"⚠️  Getting continuation...")
                
                # Get the last part of the response for context
                last_context = response_text[-1000:] if len(response_text) > 1000 else response_text
                
                # Count how many products we have so far
                products_count = response_text.count('"product_service_name"') or response_text.count('"products": [')
                
                continuation_prompt = f"""The previous JSON response was truncated at approximately {len(response_text)} characters. 
                
Last part of previous response for context:
{last_context}

Please continue the JSON from where it was cut off. IMPORTANT:
1. Continue the "products" array - do NOT repeat products already included
2. Complete any incomplete product objects from the previous response
3. Include ALL remaining products that were NOT included yet
4. After all products, add the "totals" object with calculated sums
5. Close the JSON properly with all closing brackets: ]}}}}

Return ONLY the continuation JSON starting from where the previous response ended. Do NOT include the opening {{ or "products": [ again - just continue from the last product object."""
                
                try:
                    continuation_response = self.client.chat.completions.create(
                        model=self.azure_openai_deployment,
                        messages=[
                            {
                                "role": "system",
                                "content": "You are a data extraction expert. Continue the JSON response from where it was truncated. Return ONLY valid JSON continuation without repeating already included data."
                            },
                            {
                                "role": "user",
                                "content": continuation_prompt
                            }
                        ],
                        max_tokens=max_tokens,
                        temperature=0.1
                    )
                    
                    continuation_text = continuation_response.choices[0].message.content
                    finish_reason = continuation_response.choices[0].finish_reason
                    all_responses.append(continuation_text)
                    
                    # Clean up the previous response to remove incomplete structures
                    response_text = response_text.rstrip()
                    
                    # Find the last complete product object by looking for closing braces
                    # We want to find where a complete product object ends (} followed by , or ])
                    last_complete_pos = len(response_text)
                    
                    # Try to find the last complete product object
                    # Look backwards for }, pattern (end of product object)
                    for i in range(len(response_text) - 2, -1, -1):
                        if response_text[i:i+2] == '},':
                            # Found end of a product object
                            last_complete_pos = i + 2
                            break
                        elif response_text[i] == '}' and i < len(response_text) - 1:
                            next_chars = response_text[i+1:].lstrip()
                            if next_chars.startswith(',') or next_chars.startswith(']'):
                                last_complete_pos = i + 1
                                break
                    
                    # Trim to the last complete position
                    if last_complete_pos < len(response_text):
                        response_text = response_text[:last_complete_pos].rstrip()
                    
                    # Remove trailing comma if present
                    response_text = response_text.rstrip(',').rstrip()
                    
                    # Clean up continuation text
                    continuation_text = continuation_text.strip()
                    
                    # Remove leading { if present
                    if continuation_text.startswith('{'):
                        continuation_text = continuation_text[1:].lstrip()
                    
                    # Remove "products": [ if present at the start
                    if continuation_text.startswith('"products"'):
                        # Find and remove "products": [ part
                        bracket_pos = continuation_text.find('[')
                        if bracket_pos != -1:
                            continuation_text = continuation_text[bracket_pos + 1:].lstrip()
                    
                    # Remove leading [ if present (from products array)
                    if continuation_text.startswith('['):
                        continuation_text = continuation_text[1:].lstrip()
                    
                    # Ensure proper comma separation before adding continuation
                    if response_text and not response_text.rstrip().endswith(',') and not response_text.rstrip().endswith('['):
                        response_text += ','
                    
                    # Add continuation
                    response_text += continuation_text
                    
                    print(f"✓ Continuation {continuation_count} received and merged ({len(continuation_text)} chars)")
                    
                    # If continuation was also truncated, continue the loop
                    if finish_reason != "length":
                        print("✓ All data received successfully")
                        break
                    
                except Exception as e:
                    print(f"⚠️  Could not get continuation {continuation_count}: {e}")
                    print("⚠️  Using partial response. Data may be incomplete.")
                    break
            
            if continuation_count > 0:
                print(f"✓ Merged {continuation_count} continuation(s). Total response length: {len(response_text):,} characters")
            
            if finish_reason == "length" and continuation_count >= max_continuations:
                print(f"⚠️  WARNING: Maximum continuations reached ({max_continuations}). Response may still be incomplete.")
            
            elapsed = time.time() - start_time
            print(f"✓ API call completed in {elapsed:.1f}s")
            return response_text
            
        except APITimeoutError:
            if retry_count < 2:
                print(f"⚠️  Timeout, retrying... ({retry_count + 1}/2)")
                time.sleep(5)
                return self._call_gpt4_api(text_content, tables_content, retry_count + 1, total_pages)
            raise Exception("API timeout after retries. Document may be too large.")
        
        except RateLimitError:
            print("⚠️  Rate limit hit, waiting 60s...")
            time.sleep(60)
            return self._call_gpt4_api(text_content, tables_content, retry_count, total_pages)
        
        except Exception as e:
            print(f"❌ API error: {e}")
            raise
    
    def _extract_json_from_response(self, response_text: str) -> Dict[str, Any]:
        """Extract JSON from API response - handles truncated responses and malformed JSON."""
        text = response_text.strip()
        
        # Remove markdown code blocks
        if "```json" in text:
            start = text.find("```json") + 7
            end = text.find("```", start)
            if end == -1:
                end = len(text)
            text = text[start:end].strip()
        elif "```" in text:
            start = text.find("```") + 3
            end = text.find("```", start)
            if end == -1:
                end = len(text)
            text = text[start:end].strip()
        
        # Find JSON start
        start = text.find("{")
        if start == -1:
            raise ValueError("No JSON found in response")
        
        text = text[start:]
        
        # Save raw response for debugging
        debug_dir = Path("data")
        debug_dir.mkdir(exist_ok=True)
        raw_response_path = debug_dir / "raw_gpt_response.json"
        try:
            with open(raw_response_path, 'w', encoding='utf-8') as f:
                f.write(text)
        except:
            pass
        
        # Try to parse directly first
        try:
            return json.loads(text)
        except json.JSONDecodeError as initial_error:
            print(f"❌ Initial JSON parse error: {initial_error}")
            print(f"❌ Response length: {len(response_text)} characters")
            print(f"❌ Attempting JSON repair...")
            
            # Try multiple repair strategies
            repaired_text = self._repair_json(text, initial_error)
            
            try:
                return json.loads(repaired_text)
            except json.JSONDecodeError as repair_error:
                print(f"❌ JSON repair failed: {repair_error}")
                print(f"❌ Attempting to extract partial data...")
                
                # Last resort: extract what we can
                return self._extract_partial_json(text)
    
    def _repair_json(self, text: str, error: json.JSONDecodeError) -> str:
        """Attempt to repair common JSON issues."""
        repaired = text
        
        # Fix common issues
        # 1. Fix missing commas between objects in arrays (more robust)
        # Look for } followed by { or } followed by " (new key)
        repaired = re.sub(r'}\s*\n\s*{', '},\n    {', repaired)
        repaired = re.sub(r'}\s*"', '}, "', repaired)
        repaired = re.sub(r']\s*"', '], "', repaired)
        repaired = re.sub(r'}\s*\[', '}, [', repaired)
        
        # 2. Fix trailing commas
        repaired = re.sub(r',\s*}', '}', repaired)
        repaired = re.sub(r',\s*]', ']', repaired)
        
        # 3. Fix missing commas after values (common cause of "Expecting ',' delimiter")
        # Pattern: value followed by " (new key) or { (new object) or [ (new array)
        repaired = re.sub(r'"\s*\n\s*"', '",\n    "', repaired)  # String to string
        repaired = re.sub(r'(\d+)\s*\n\s*"', r'\1,\n    "', repaired)  # Number to string
        repaired = re.sub(r'(true|false|null)\s*\n\s*"', r'\1,\n    "', repaired)  # Boolean/null to string
        repaired = re.sub(r'}\s*\n\s*"', '},\n    "', repaired)  # Object to string
        repaired = re.sub(r']\s*\n\s*"', '],\n    "', repaired)  # Array to string
        
        # 4. Fix unclosed strings (common issue)
        # Find the error position and try to fix it
        if hasattr(error, 'pos'):
            error_pos = error.pos
            # Check if we're in the middle of a string
            before_error = repaired[:error_pos]
            after_error = repaired[error_pos:]
            
            # Count unescaped quotes before error
            quote_count = before_error.count('"') - before_error.count('\\"')
            if quote_count % 2 == 1:  # Odd number means unclosed string
                # Try to close the string
                # Find the next safe place to insert a quote
                next_comma = after_error.find(',')
                next_brace = after_error.find('}')
                next_bracket = after_error.find(']')
                next_newline = after_error.find('\n')
                
                safe_pos = min([p for p in [next_comma, next_brace, next_bracket, next_newline] if p != -1], default=0)
                if safe_pos > 0:
                    repaired = before_error + '"' + after_error[:safe_pos] + after_error[safe_pos:]
        
        # 4. Check if JSON is incomplete (missing closing braces/brackets)
        open_braces = repaired.count('{')
        close_braces = repaired.count('}')
        open_brackets = repaired.count('[')
        close_brackets = repaired.count(']')
        
        if open_braces > close_braces or open_brackets > close_brackets:
            print("⚠️  WARNING: JSON appears incomplete, attempting to fix structure...")
            
            # Try to extract and fix products array
            if '"products": [' in repaired:
                products_start = repaired.find('"products": [')
                products_section = repaired[products_start:]
                
                # Try to find valid product entries
                # Look for complete product objects (starting with { and ending with })
                product_pattern = r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}'
                products = re.findall(product_pattern, products_section)
                
                if products:
                    # Rebuild products array with valid entries
                    products_json = ',\n    '.join(products)
                    # Replace the products section
                    products_end = products_section.find(']')
                    if products_end == -1:
                        # No closing bracket, add one
                        new_products_section = f'"products": [\n    {products_json}\n  ]'
                        repaired = repaired[:products_start] + new_products_section
                    else:
                        # Replace existing products
                        new_products_section = f'"products": [\n    {products_json}\n  ]'
                        repaired = repaired[:products_start] + new_products_section + products_section[products_end+1:]
            
            # Close totals if missing
            if '"totals": {' not in repaired:
                # Check if we need to add it before closing braces
                if '"products":' in repaired:
                    products_end = repaired.rfind(']')
                    if products_end != -1:
                        totals_section = ',\n  "totals": {\n    "total_quantity": 0,\n    "total_gross_quantity": 0,\n    "total_net_weight": 0,\n    "total_gross_weight": 0\n  }'
                        repaired = repaired[:products_end+1] + totals_section + repaired[products_end+1:]
            
            # Close any remaining open structures
            missing_braces = open_braces - close_braces
            missing_brackets = open_brackets - close_brackets
            if missing_braces > 0:
                repaired += '}' * missing_braces
            if missing_brackets > 0:
                repaired += ']' * missing_brackets
        
        return repaired
    
    def _extract_partial_json(self, text: str) -> Dict[str, Any]:
        """Extract partial JSON data when full parsing fails."""
        result = {
            "document_metadata": {},
            "products": [],
            "totals": {
                "total_quantity": 0,
                "total_gross_quantity": 0,
                "total_net_weight": 0,
                "total_gross_weight": 0
            }
        }
        
        # Try to extract metadata
        try:
            # Find metadata section - handle nested braces
            metadata_start = text.find('"document_metadata"')
            if metadata_start != -1:
                # Find the opening brace
                brace_start = text.find('{', metadata_start)
                if brace_start != -1:
                    # Count braces to find the end
                    brace_count = 0
                    metadata_end = brace_start
                    for i in range(brace_start, min(brace_start + 5000, len(text))):
                        if text[i] == '{':
                            brace_count += 1
                        elif text[i] == '}':
                            brace_count -= 1
                            if brace_count == 0:
                                metadata_end = i + 1
                                break
                    
                    if metadata_end > brace_start:
                        metadata_str = text[brace_start:metadata_end]
                        try:
                            result["document_metadata"] = json.loads(metadata_str)
                        except:
                            # Extract metadata fields manually
                            metadata_fields = ['document_type', 'document_number', 'date_of_issue']
                            for field in metadata_fields:
                                pattern = f'"{field}"\\s*:\\s*"([^"]*)"'
                                match = re.search(pattern, metadata_str)
                                if match:
                                    if 'document_metadata' not in result or not isinstance(result['document_metadata'], dict):
                                        result['document_metadata'] = {}
                                    result['document_metadata'][field] = match.group(1)
        except Exception as e:
            print(f"⚠️  Could not extract metadata: {e}")
        
        # Try to extract products - more robust approach
        try:
            products_start = text.find('"products": [')
            if products_start != -1:
                # Find the products array content
                bracket_start = text.find('[', products_start)
                if bracket_start != -1:
                    # Extract all content between [ and ]
                    # We'll look for individual product objects
                    products_section = text[bracket_start:]
                    
                    # Find all product objects - look for opening braces
                    # This is more robust than regex
                    i = 0
                    while i < len(products_section):
                        # Find next opening brace
                        obj_start = products_section.find('{', i)
                        if obj_start == -1:
                            break
                        
                        # Find matching closing brace
                        brace_count = 0
                        obj_end = obj_start
                        in_string = False
                        escape_next = False
                        
                        for j in range(obj_start, len(products_section)):
                            char = products_section[j]
                            
                            if escape_next:
                                escape_next = False
                                continue
                            
                            if char == '\\':
                                escape_next = True
                                continue
                            
                            if char == '"' and not escape_next:
                                in_string = not in_string
                                continue
                            
                            if not in_string:
                                if char == '{':
                                    brace_count += 1
                                elif char == '}':
                                    brace_count -= 1
                                    if brace_count == 0:
                                        obj_end = j + 1
                                        break
                        
                        if obj_end > obj_start:
                            product_str = products_section[obj_start:obj_end]
                            
                            # Try to parse the product
                            try:
                                # Fix common issues
                                fixed_product = product_str
                                # Remove trailing commas
                                fixed_product = re.sub(r',\s*}', '}', fixed_product)
                                # Fix unclosed strings (basic)
                                fixed_product = re.sub(r':\s*"([^"]*?)(?=\s*[,}])', r': "\1"', fixed_product)
                                
                                product = json.loads(fixed_product)
                                result["products"].append(product)
                            except:
                                # Extract fields manually using regex
                                product = {}
                                # Extract string fields
                                string_fields = ['product_service_name', 'article', 'batch_number', 'order_number', 
                                               'ean', 'fabric_colors', 'site']
                                for field in string_fields:
                                    pattern = f'"{field}"\\s*:\\s*"([^"]*)"'
                                    match = re.search(pattern, product_str)
                                    if match:
                                        product[field] = match.group(1)
                                
                                # Extract numeric fields
                                numeric_fields = ['quantity', 'gross_quantity', 'net_weight', 'gross_weight']
                                for field in numeric_fields:
                                    pattern = f'"{field}"\\s*:\\s*([0-9.]+)'
                                    match = re.search(pattern, product_str)
                                    if match:
                                        try:
                                            product[field] = float(match.group(1))
                                        except:
                                            pass
                                
                                if product:
                                    result["products"].append(product)
                            
                            i = obj_end
                        else:
                            break
        except Exception as e:
            print(f"⚠️  Could not extract products: {e}")
        
        # Calculate totals from extracted products
        if result["products"]:
            result["totals"]["total_quantity"] = sum(p.get("quantity", 0) for p in result["products"])
            result["totals"]["total_gross_quantity"] = sum(p.get("gross_quantity", 0) for p in result["products"])
            result["totals"]["total_net_weight"] = sum(p.get("net_weight", 0) for p in result["products"])
            result["totals"]["total_gross_weight"] = sum(p.get("gross_weight", 0) for p in result["products"])
            print(f"✓ Extracted {len(result['products'])} product(s) from partial JSON")
        else:
            print("⚠️  WARNING: Could not extract any products from malformed JSON")
        
        return result
    
    def _split_combined_fields(self, data: Any) -> Any:
        """Recursively split combined fields into separate key-value pairs."""
        if isinstance(data, dict):
            result = {}
            for key, value in data.items():
                if isinstance(value, list):
                    result[key] = [self._split_combined_fields(item) for item in value]
                elif isinstance(value, dict):
                    result[key] = self._split_combined_fields(value)
                elif isinstance(value, str) and any(k in key.lower() for k in ['bnumber', 'batch', 'reference', 'ean']):
                    if ',' in value and ('order' in value.lower() or 'ean' in value.lower()):
                        split_data = self._parse_combined_field(value)
                        result.update(split_data)
                    result[key] = value
                else:
                    result[key] = value
            return result
        elif isinstance(data, list):
            return [self._split_combined_fields(item) for item in data]
        else:
            return data
    
    def _parse_combined_field(self, field_value: str) -> Dict[str, str]:
        """Parse combined field into separate components."""
        result = {}
        if not isinstance(field_value, str):
            return result
        
        # Extract Order number
        order_match = re.search(r'Order:\s*(\d+)', field_value, re.IGNORECASE)
        if order_match:
            result['order_number'] = order_match.group(1)
        
        # Extract EAN
        ean_match = re.search(r'EAN:\s*([A-Z0-9]+)', field_value, re.IGNORECASE)
        if ean_match:
            result['ean'] = ean_match.group(1)
        
        # Extract batch number
        batch_match = re.search(r'([A-Z]+\/[0-9\/]+\/[A-Z]+)', field_value)
        if batch_match:
            result['batch_number'] = batch_match.group(1)
        else:
            parts = [p.strip() for p in field_value.split(',')]
            if len(parts) > 1:
                first_part = parts[0]
                if '/' in first_part or re.match(r'^[A-Z0-9\/\-]+$', first_part):
                    result['batch_number'] = first_part
        
        return result
    
    def convert_to_json(
        self,
        file_path: str,
        output_path: Optional[str] = None,
        custom_prompt: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Convert document to JSON.
        
        Args:
            file_path: Path to document file (PDF, images, Excel, etc.)
            output_path: Optional output JSON path
            custom_prompt: Optional custom extraction prompt
            
        Returns:
            Extracted data as dictionary
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        original_file = str(file_path)
        converted_file = None
        cleanup_converted = False
        
        # Check if it's an Excel file - handle directly without PDF conversion
        if self._is_excel_file(original_file):
            print(f"  Detected Excel file - extracting data directly (no PDF conversion)...")
            try:
                # Extract data directly from Excel
                extraction_result = self._extract_excel_data(original_file)
                text_content = extraction_result['text']
                tables = extraction_result['tables']
                total_pages = extraction_result.get('total_pages', 1)
                
                if not text_content.strip():
                    raise Exception("No data could be extracted from the Excel file.")
            except Exception as e:
                print(f"❌ Error extracting Excel data: {e}")
                raise
        else:
            # For non-Excel files, use Azure Document Intelligence
            # Convert unsupported formats to PDF if needed
            if not self._is_supported_format(original_file):
                print(f"  File format not directly supported, converting to PDF...")
                # Note: Excel files are now handled above, so this won't be called for Excel
                raise ValueError(
                    f"Unsupported file format: {Path(original_file).suffix}\n"
                    f"Supported formats: PDF, PNG, JPEG, TIFF, BMP, DOCX, DOC, PPTX, PPT, XLSX, XLS"
                )
            
            try:
                # Extract text from document using Azure Document Intelligence
                extraction_result = self._extract_document_with_azure(original_file)
                text_content = extraction_result['text']
                tables = extraction_result['tables']
                total_pages = extraction_result.get('total_pages', 1)
                
                if not text_content.strip():
                    raise Exception("No text could be extracted from the document. It may be a scanned/image-only document.")
            except Exception as e:
                print(f"❌ Error extracting document: {e}")
                raise
            
        # Common processing for both Excel and other documents
        try:
            # Count total table rows for validation
            total_table_rows = sum(table_info.get('row_count', len(table_info.get('data', []))) for table_info in tables)
            if tables:
                print(f"📊 Table extraction summary: {len(tables)} table(s) with {total_table_rows} total row(s)")
                # Estimate data rows (subtract header rows - typically 1 per table)
                estimated_data_rows = total_table_rows - len(tables)
                print(f"   Expected products in output: ~{estimated_data_rows} (subtracting header rows)")
            
            # Verify all pages/sheets are included in text
            page_numbers_in_text = set()
            for line in text_content.split('\n'):
                if line.startswith('=== Page ') or line.startswith('=== Sheet '):
                    try:
                        # Extract page/sheet number
                        if 'Page ' in line:
                            page_num = int(line.split('Page ')[1].split(' ===')[0])
                        elif 'Sheet ' in line:
                            page_num = int(line.split('Sheet ')[1].split(':')[0])
                        else:
                            continue
                        page_numbers_in_text.add(page_num)
                    except:
                        pass
            
            if page_numbers_in_text:
                print(f"📄 Pages/Sheets included in extraction: {sorted(page_numbers_in_text)}")
                if len(page_numbers_in_text) < total_pages:
                    missing_pages = set(range(1, total_pages + 1)) - page_numbers_in_text
                    print(f"⚠️  WARNING: Missing pages/sheets: {sorted(missing_pages)}")
            
            # Determine if we need to chunk the data
            total_table_rows = sum(table_info.get('row_count', len(table_info.get('data', []))) for table_info in tables)
            estimated_data_rows = total_table_rows - len(tables) if tables else 0
            
            # Use chunking if we have more than 80 data rows (to avoid token limit issues)
            # Reduced threshold to ensure we chunk earlier and use smaller chunks
            use_chunking = estimated_data_rows > 80 and tables
            
            if use_chunking:
                print(f"\n📦 Large dataset detected ({estimated_data_rows} data rows). Processing in chunks to avoid token limits...")
                # Reduced chunk size to 100 rows to ensure each chunk stays well under token limit
                chunks = self._chunk_tables(tables, chunk_size=100)
                print(f"   Split into {len(chunks)} chunk(s) of ~100 rows each")
                
                all_products = []
                document_metadata = None
                chunk_results = []
                
                # Process each chunk
                for chunk_idx, (start_row, end_row) in enumerate(chunks, 1):
                    print(f"\n  Processing chunk {chunk_idx}/{len(chunks)} (rows {start_row + 1} to {end_row if end_row else 'end'})...")
                    
                    # Format tables for this chunk
                    tables_content = self._format_tables_for_prompt(tables, start_row=start_row, end_row=end_row) if tables else ""
                    
                    # Use custom prompt if provided, otherwise use default
                    if custom_prompt:
                        prompt_text = custom_prompt
                    else:
                        prompt_text = text_content
                    
                    # Call GPT-4 to structure the data for this chunk
                    response_text = self._call_gpt4_api(prompt_text, tables_content, total_pages=total_pages, chunk_info=f"chunk {chunk_idx}/{len(chunks)}")
                    
                    # Extract JSON from response
                    try:
                        chunk_json_data = self._extract_json_from_response(response_text)
                        
                        # Extract metadata from first chunk
                        if chunk_idx == 1 and 'document_metadata' in chunk_json_data:
                            document_metadata = chunk_json_data['document_metadata']
                        
                        # Extract products from this chunk
                        if 'products' in chunk_json_data and chunk_json_data['products']:
                            chunk_products = chunk_json_data['products']
                            all_products.extend(chunk_products)
                            print(f"    ✓ Extracted {len(chunk_products)} product(s) from chunk {chunk_idx}")
                        
                        chunk_results.append(chunk_json_data)
                        
                    except Exception as e:
                        print(f"    ❌ Error processing chunk {chunk_idx}: {e}")
                        print(f"    ⚠️  Skipping chunk {chunk_idx}, data may be incomplete")
                        continue
                
                # Merge all chunks into final data
                print(f"\n  Merging {len(chunks)} chunk(s)...")
                
                # Calculate totals from all products
                totals = {
                    "total_quantity": sum(p.get("quantity", 0) for p in all_products),
                    "total_gross_quantity": sum(p.get("gross_quantity", 0) for p in all_products),
                    "total_net_weight": sum(p.get("net_weight", 0) for p in all_products),
                    "total_gross_weight": sum(p.get("gross_weight", 0) for p in all_products)
                }
                
                # Build final data structure
                final_data = {
                    "document_metadata": document_metadata or {},
                    "products": all_products,
                    "totals": totals
                }
                
                print(f"  ✓ Merged {len(all_products)} total product(s)")
                
                # Post-process to split combined fields
                print("Processing combined fields...")
                final_data = self._split_combined_fields(final_data)
                
            else:
                # Process normally without chunking
                # Format tables for prompt
                tables_content = self._format_tables_for_prompt(tables) if tables else ""
                
                # Use custom prompt if provided, otherwise use default
                if custom_prompt:
                    prompt_text = custom_prompt
                else:
                    prompt_text = text_content
                
                # Call GPT-4 to structure the data
                response_text = self._call_gpt4_api(prompt_text, tables_content, total_pages=total_pages)
                
                # Extract JSON from response
                try:
                    json_data = self._extract_json_from_response(response_text)
                    
                    # Check if we got all expected products
                    products_extracted = len(json_data.get('products', []))
                    if estimated_data_rows > 0 and products_extracted < estimated_data_rows * 0.8:
                        # We're missing more than 20% of products - likely truncated
                        print(f"\n⚠️  WARNING: Only extracted {products_extracted} products but expected ~{estimated_data_rows}")
                        print(f"⚠️  Response may have been truncated. Retrying with chunking...")
                        
                        # Retry with chunking
                        use_chunking = True
                        chunks = self._chunk_tables(tables, chunk_size=100)
                        print(f"   Split into {len(chunks)} chunk(s) of ~100 rows each")
                        
                        all_products = []
                        document_metadata = json_data.get('document_metadata', {})
                        
                        # Process each chunk
                        for chunk_idx, (start_row, end_row) in enumerate(chunks, 1):
                            print(f"\n  Processing chunk {chunk_idx}/{len(chunks)} (rows {start_row + 1} to {end_row if end_row else 'end'})...")
                            
                            # Format tables for this chunk
                            chunk_tables_content = self._format_tables_for_prompt(tables, start_row=start_row, end_row=end_row) if tables else ""
                            
                            # Call GPT-4 to structure the data for this chunk
                            chunk_response = self._call_gpt4_api(prompt_text, chunk_tables_content, total_pages=total_pages, chunk_info=f"chunk {chunk_idx}/{len(chunks)}")
                            
                            # Extract JSON from response
                            try:
                                chunk_json_data = self._extract_json_from_response(chunk_response)
                                
                                # Extract products from this chunk
                                if 'products' in chunk_json_data and chunk_json_data['products']:
                                    chunk_products = chunk_json_data['products']
                                    all_products.extend(chunk_products)
                                    print(f"    ✓ Extracted {len(chunk_products)} product(s) from chunk {chunk_idx}")
                                
                            except Exception as e:
                                print(f"    ❌ Error processing chunk {chunk_idx}: {e}")
                                print(f"    ⚠️  Skipping chunk {chunk_idx}, data may be incomplete")
                                continue
                        
                        # Merge all chunks
                        print(f"\n  Merging {len(chunks)} chunk(s)...")
                        
                        # Calculate totals
                        totals = {
                            "total_quantity": sum(p.get("quantity", 0) for p in all_products),
                            "total_gross_quantity": sum(p.get("gross_quantity", 0) for p in all_products),
                            "total_net_weight": sum(p.get("net_weight", 0) for p in all_products),
                            "total_gross_weight": sum(p.get("gross_weight", 0) for p in all_products)
                        }
                        
                        # Build final data structure
                        final_data = {
                            "document_metadata": document_metadata,
                            "products": all_products,
                            "totals": totals
                        }
                        
                        print(f"  ✓ Merged {len(all_products)} total product(s)")
                    else:
                        # Normal processing - no chunking needed
                        final_data = json_data
                    
                    # Post-process to split combined fields
                    print("Processing combined fields...")
                    final_data = self._split_combined_fields(final_data)
                except Exception as e:
                    print(f"❌ Error processing JSON: {e}")
                    print(f"⚠️  The response may have been truncated. Check if the document is too large.")
                    # Cleanup converted file on error
                    if cleanup_converted and converted_file and os.path.exists(converted_file):
                        try:
                            os.remove(converted_file)
                        except:
                            pass
                    raise
            
            # Save to file - default to data/pl_data.json
            if not output_path:
                data_dir = file_path.parent / "data"
                data_dir.mkdir(parents=True, exist_ok=True)
                output_path = data_dir / "pl_data.json"
            else:
                output_path = Path(output_path)
                output_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(final_data, f, indent=2, ensure_ascii=False)
            
            print(f"\n✓ JSON saved to: {output_path}")
            
            # Show summary and validate
            if isinstance(final_data, dict):
                if 'products' in final_data:
                    product_count = len(final_data['products'])
                    print(f"✓ Extracted {product_count} product(s)")
                    if tables:
                        estimated_data_rows = total_table_rows - len(tables)
                        if product_count < estimated_data_rows * 0.9:  # Allow 10% tolerance
                            print(f"⚠️  WARNING: Expected ~{estimated_data_rows} products but got {product_count}. Some rows may be missing!")
                        else:
                            print(f"✓ Product count looks good (expected ~{estimated_data_rows}, got {product_count})")
                elif 'items' in final_data:
                    print(f"✓ Extracted {len(final_data['items'])} item(s)")
            
            # Cleanup converted file if created
            if cleanup_converted and converted_file and os.path.exists(converted_file):
                try:
                    os.remove(converted_file)
                    print(f"  ✓ Cleaned up converted file")
                except Exception as e:
                    print(f"  ⚠️  Could not delete converted file: {e}")
            
            return final_data
            
        except Exception as e:
            # Cleanup converted file on error
            if cleanup_converted and converted_file and os.path.exists(converted_file):
                try:
                    os.remove(converted_file)
                except:
                    pass
            raise

